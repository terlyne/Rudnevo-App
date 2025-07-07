import os
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Any
import jinja2
from pathlib import Path
import json


class ScheduleTemplateGenerator:
    """Генератор шаблонов расписаний из Excel файлов"""
    
    def __init__(self, templates_dir: str = "templates"):
        # Используем абсолютный путь относительно backend
        base_dir = Path(__file__).parent.parent.parent
        self.templates_dir = base_dir / templates_dir
        self.templates_dir.mkdir(exist_ok=True)
        
        # Настройка Jinja2
        template_loader = jinja2.FileSystemLoader(searchpath=str(self.templates_dir))
        self.template_env = jinja2.Environment(loader=template_loader)
    
    def process_excel_to_templates(self, file_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """Обработка Excel файла и генерация шаблонов"""
        wb = load_workbook(file_path)
        result = []
        
        for sheet_name in wb.sheetnames:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df = df.fillna("")
            
            # Генерируем HTML шаблон
            html_content = self.generate_html_template(sheet_name, df)
            
            # Сохраняем шаблон
            template_path = self.templates_dir / f"{sheet_name}.html"
            with open(template_path, "w", encoding="utf-8") as f:
                f.write(html_content)
            
            result.append({
                "college": sheet_name,
                "schedule": df.to_dict(orient="records"),
                "template_url": f"/api/v1/schedule/templates/{sheet_name}.html",
                "file_path": str(template_path)
            })
        
        return {"status": "success", "data": result}
    
    def generate_html_template(self, college_name: str, df: pd.DataFrame) -> str:
        """Генерация HTML шаблона с поддержкой стилей"""
        template_content = self._get_base_template()
        template = self.template_env.from_string(template_content)
        
        return template.render(
            college=college_name,
            headers=df.columns.tolist(),
            rows=df.values.tolist()
        )
    
    def _get_base_template(self) -> str:
        """Базовый HTML шаблон для расписаний"""
        return """
<!DOCTYPE html>
<html>
<head>
    <title>{{ college }} - Расписание</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f7fa;
        }

        .schedule-container {
            max-width: 1000px;
            margin: 0 auto;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
            padding: 20px;
        }

        h1 {
            color: #2c3e50;
            text-align: center;
            margin-bottom: 20px;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }

        th, td {
            padding: 12px 15px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }

        th {
            background-color: #3A86FF;
            color: white;
            font-weight: 600;
        }

        tr:nth-child(even) {
            background-color: #f8f9fa;
        }

        tr:hover {
            background-color: #e9f1ff;
        }

        .empty-cell {
            color: #999;
            font-style: italic;
        }

        @media (max-width: 768px) {
            .schedule-container {
                padding: 10px;
                margin: 10px;
            }
            
            th, td {
                padding: 8px 10px;
                font-size: 14px;
            }
            
            h1 {
                font-size: 1.5rem;
            }
        }
    </style>
</head>

<body>
    <div class="schedule-container">
        <h1>{{ college }} - Расписание</h1>

        <table>
            <thead>
                <tr>
                    {% for header in headers %}
                    <th>{{ header }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for row in rows %}
                <tr>
                    {% for cell in row %}
                    <td class="{{ 'empty-cell' if cell == '' else '' }}">
                        {{ cell if cell != '' else 'Нет данных' }}
                    </td>
                    {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
</body>
</html>
        """
    
    def get_template_content(self, template_name: str) -> str:
        """Получить содержимое шаблона по имени"""
        template_path = self.templates_dir / f"{template_name}.html"
        if template_path.exists():
            with open(template_path, "r", encoding="utf-8") as f:
                return f.read()
        return ""


# Создаем глобальный экземпляр генератора
schedule_generator = ScheduleTemplateGenerator() 


async def process_excel_schedule(file_path: str, db_session=None) -> List[Dict[str, Any]]:
    """
    Обрабатывает Excel файл с расписаниями и возвращает список JSON данных
    для каждого колледжа. Более гибкая обработка - использует первый лист.
    """
    try:
        # Читаем Excel файл через контекстный менеджер
        with pd.ExcelFile(file_path) as excel_file:
            results = []
            # Получаем список активных колледжей из базы данных
            available_colleges = set()
            colleges_by_name = dict()
            if db_session:
                try:
                    from crud import college as college_crud
                    colleges = await college_crud.get_colleges_list(db_session)
                    available_colleges = {college.name for college in colleges}
                    colleges_by_name = {college.name: college.id for college in colleges}
                except Exception as e:
                    print(f"Ошибка при получении списка колледжей: {str(e)}")

            print(f"=== DEBUG: Обработка Excel файла ===")
            print(f"Доступные колледжи в БД: {available_colleges}")
            print(f"Листы в Excel файле: {excel_file.sheet_names}")

            # Используем только первый лист для гибкости
            if not excel_file.sheet_names:
                raise Exception("Excel файл не содержит листов")

            sheet_name = excel_file.sheet_names[0]
            print(f"\n--- Обрабатываем первый лист: '{sheet_name}' ---")

            try:
                # Читаем данные из листа через ExcelFile
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                # Обрабатываем данные и создаем JSON структуру
                schedule_data = process_sheet_data(df, sheet_name)

                if schedule_data:
                    print(f"Успешно созданы данные для листа '{sheet_name}'")

                    # Используем название листа как название колледжа
                    college_name = sheet_name
                    print(f"Используем название листа как название колледжа: '{college_name}'")

                    # Проверяем, есть ли такой колледж в БД
                    college_id = None
                    if available_colleges and college_name in available_colleges:
                        college_id = colleges_by_name.get(college_name)
                    else:
                        print(f"ВНИМАНИЕ: Колледж '{college_name}' не найден в базе данных!")
                        print(f"Доступные колледжи: {available_colleges}")
                        # Используем первый доступный колледж
                        if available_colleges:
                            college_name = list(available_colleges)[0]
                            college_id = colleges_by_name.get(college_name)
                            print(f"Используем первый доступный колледж: '{college_name}' (id={college_id})")
                        else:
                            print(f"ВНИМАНИЕ: Нет колледжей в базе данных!")
                            # Создаем временное название
                            college_name = f"Колледж_{sheet_name}"
                            college_id = None
                            print(f"Создаем временное название: '{college_name}' (id=None)")

                    results.append({
                        "college_id": college_id,
                        "college_name": college_name,
                        "schedule_data": schedule_data
                    })
                else:
                    print(f"Не удалось создать данные для листа '{sheet_name}' - пустой результат")

            except Exception as e:
                print(f"Ошибка при обработке листа '{sheet_name}': {str(e)}")
                raise

            print(f"=== ИТОГОВЫЙ РЕЗУЛЬТАТ ===")
            print(f"Обработано колледжей: {len(results)}")
            for result in results:
                print(f"  - {result['college_name']} (id={result['college_id']}): {len(result['schedule_data']['schedule'])} записей")

            return results
    except Exception as e:
        print(f"Ошибка при обработке Excel файла: {str(e)}")
        raise


def process_sheet_data(df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
    """
    Обрабатывает данные из листа Excel и создает структурированный JSON
    """
    print(f"  Обрабатываем данные для листа '{sheet_name}'")
    print(f"  Исходный DataFrame: {df.shape[0]} строк, {df.shape[1]} колонок")
    print(f"  Заголовки: {df.columns.tolist()}")
    
    # Очищаем DataFrame
    df = df.dropna(how='all').dropna(axis=1, how='all')
    print(f"  После очистки: {df.shape[0]} строк, {df.shape[1]} колонок")
    
    # Если DataFrame пустой, возвращаем None
    if df.empty:
        print(f"  DataFrame пустой после очистки")
        return None
    
    # Определяем структуру расписания
    schedule_structure = {
        "college_name": sheet_name,
        "last_updated": pd.Timestamp.now().isoformat(),
        "schedule": []
    }
    
    # Обрабатываем данные в зависимости от структуры
    if len(df.columns) >= 2:
        # Предполагаем, что первая колонка - группа, остальные - данные расписания
        schedule_structure["schedule"] = process_standard_schedule(df)
    else:
        # Простая структура - все данные в одной колонке
        schedule_structure["schedule"] = process_simple_schedule(df)
    
    print(f"  Создана структура с {len(schedule_structure['schedule'])} записями")
    return schedule_structure


def process_standard_schedule(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Обрабатывает стандартную структуру расписания и создает список записей
    с конкретными полями для отображения
    """
    schedule_entries = []
    
    # Получаем заголовки
    headers = df.columns.tolist()
    
    # Отладочная информация
    print(f"    === DEBUG: process_standard_schedule ===")
    print(f"    Заголовки колонок: {headers}")
    print(f"    Количество строк: {len(df)}")
    print(f"    Первые 3 строки:")
    print(df.head(3))
    
    # Обрабатываем каждую строку
    for index, row in df.iterrows():
        # Первая колонка - название группы
        group_name = str(row.iloc[0]) if not pd.isna(row.iloc[0]) else f"Группа_{index+1}"
        
        # Ищем данные в остальных колонках строго по заголовкам
        discipline = None
        room = None
        shift = None
        auditory = None
        dates = None
        
        # Проходим по всем колонкам и ищем нужные данные строго по заголовкам
        for i, header in enumerate(headers[1:], 1):
            cell_value = row.iloc[i]
            if not pd.isna(cell_value):
                cell_str = str(cell_value).strip()
                
                # Определяем тип данных строго по заголовку колонки
                header_lower = str(header).lower()
                
                print(f"Строка {index}, колонка '{header}' (индекс {i}): '{cell_str}'")
                
                # Строгое соответствие заголовкам
                if 'дисциплина' in header_lower:
                    discipline = cell_str
                    print(f"  -> Найдена дисциплина: {discipline}")
                elif 'название аудитории' in header_lower:
                    auditory = cell_str
                    print(f"  -> Найдена аудитория: {auditory}")
                elif 'номер кабинета' in header_lower:
                    room = cell_str
                    print(f"  -> Найден кабинет: {room}")
                elif 'смена' in header_lower:
                    shift = cell_str
                    print(f"  -> Найдена смена: {shift}")
                elif 'даты' in header_lower:
                    dates = cell_str
                    print(f"  -> Найдены даты: {dates}")
        
        # Создаем запись расписания
        if group_name:
            entry = {
                "group": group_name,
                "discipline": discipline or "",
                "room": room or "",
                "shift": shift or "",
                "auditory": auditory or "",
                "dates": dates or ""
            }
            schedule_entries.append(entry)
            print(f"Создана запись: {entry}")
    
    print(f"    Всего создано записей: {len(schedule_entries)}")
    return schedule_entries


def process_simple_schedule(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Обрабатывает простую структуру расписания
    """
    schedule_entries = []
    
    # Просто преобразуем DataFrame в список записей
    for index, row in df.iterrows():
        entry = {
            "group": f"Группа_{index+1}",
            "discipline": "",
            "room": "",
            "shift": "",
            "auditory": "",
            "dates": ""
        }
        
        # Заполняем данные из колонок
        for col in df.columns:
            if not pd.isna(row[col]):
                cell_str = str(row[col]).strip()
                col_lower = str(col).lower()
                
                # Сначала проверяем более специфичные условия
                if any(word in col_lower for word in ['аудитория', 'зал']):
                    entry["auditory"] = cell_str
                elif any(word in col_lower for word in ['кабинет', 'комната', 'номер']):
                    entry["room"] = cell_str
                elif 'смена' in col_lower:
                    entry["shift"] = cell_str
                elif any(word in col_lower for word in ['даты', 'дата', 'период']):
                    entry["dates"] = cell_str
                elif any(word in col_lower for word in ['дисциплина', 'предмет']):
                    entry["discipline"] = cell_str
                else:
                    # Если не определили тип, считаем это дисциплиной
                    if not entry["discipline"]:
                        entry["discipline"] = cell_str
        
        if entry["discipline"] or entry["room"] or entry["shift"] or entry["auditory"] or entry["dates"]:
            schedule_entries.append(entry)
    
    return schedule_entries


def validate_schedule_data(schedule_data: Dict[str, Any]) -> bool:
    """
    Проверяет корректность данных расписания
    """
    required_fields = ["college_name", "schedule"]
    
    for field in required_fields:
        if field not in schedule_data:
            return False
    
    if not isinstance(schedule_data["college_name"], str):
        return False
    
    if not isinstance(schedule_data["schedule"], list):
        return False
    
    return True


def format_schedule_for_display(schedule_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Форматирует данные расписания для отображения
    """
    schedule_list = schedule_data.get("schedule", [])
    formatted = {
        "college_name": schedule_data.get("college_name", ""),
        "last_updated": schedule_data.get("last_updated", ""),
        "schedule_summary": {
            "total_entries": len(schedule_list),
            "groups": [entry.get("group", "") for entry in schedule_list if isinstance(entry, dict)]
        },
        "schedule": schedule_list
    }
    
    return formatted 